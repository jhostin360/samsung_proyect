import locale
import os
import time
from tkinter import *
from tkinter import ttk
import tkinter as tk
from tkinter import Canvas
import openpyxl
import util.generic as utl
from tkinter import LabelFrame, PhotoImage, Label, Frame, messagebox
from Conexion import *
#from ventana_nuevo_profesor import ventana_register
import main
import datetime
from datetime import date


def ventana_admin():

    def cargar_datos():

        cursor = connection.cursor()

        # Consulta para obtener profesores y cantidad de estudiantes
        query = """
       SELECT p.nombre, p.apellido, p.usuario, p.contrasena, COUNT(a.id_alumnos) AS cantidad_estudiantes
    FROM profesores p
    LEFT JOIN alumnos a ON p.id_profesores = a.fk_id_profesores
    GROUP BY p.id_profesores, p.nombre, p.apellido, p.usuario, p.contrasena
        """
        cursor.execute(query)
        profesores = cursor.fetchall()
        obtener_cantidad_profesores()
        # Borrar los registros actuales de la tabla
        for item in tree.get_children():
            tree.delete(item)

        for profesor in profesores:
            tree.insert("", "end", values=(profesor[0], profesor[1], profesor[2], profesor[3], profesor[4],))

    def cargar_datos_alumnos():

        cursor = connection.cursor()

        # Consulta para obtener alumnos y su profesor
        query = """
        SELECT a.nombre, a.apellido, a.usuario, a.contrasena, p.nombre AS nombre_profesor, p.apellido AS apellido_profesor
        FROM alumnos a
        LEFT JOIN profesores p ON a.fk_id_profesores = p.id_profesores
        """
        cursor.execute(query)
        alumnos = cursor.fetchall()

        for alumno in alumnos:
            tree_alumnos.insert("", "end", values=(alumno[0], alumno[1], alumno[2], alumno[3], alumno[4] + " " + alumno[5]))

    def cargar_datos_calificaciones():
        cursor = connection.cursor()

        # Consulta para obtener calificaciones y el promedio calculado en la base de datos
        query = """
        SELECT a.nombre, a.apellido, c.Primer_examen, c.Segundo_examen, c.Tercer_examen, c.Examen_final, 
            (c.Primer_examen + c.Segundo_examen + c.Tercer_examen + c.Examen_final) / 4 as Promedio
        FROM calificaciones c
        INNER JOIN alumnos a ON c.id_alumnos = a.id_alumnos
        """
        cursor.execute(query)
        calificaciones = cursor.fetchall()

        for calificacion in calificaciones:
            nombre_completo = calificacion[0] + " " + calificacion[1]  # Concatenar nombre y apellido
            primer_examen = calificacion[2]
            segundo_examen = calificacion[3]
            tercer_examen = calificacion[4]
            examen_final = calificacion[5]
            promedio = calificacion[6]
            
            # Insertar en la tabla
            tree_calificaciones.insert("", "end", values=(nombre_completo, primer_examen, segundo_examen, tercer_examen, examen_final, promedio))

    def exportar_a_excel_profesores():
        # Crear un nuevo libro de trabajo de Excel
        libro = openpyxl.Workbook()
        hoja = libro.active

        # Agregar encabezados de columna
        encabezados = ["Nombre", "Apellido", "Usuario", "Contraseña", "Cantidad Estudiantes"]
        hoja.append(encabezados)

        # Obtener datos de la tabla de profesores
        datos = []
        for item in tree.get_children():
            valores = tree.item(item, 'values')
            datos.append(valores)

        # Agregar los datos a la hoja de Excel
        for dato in datos:
            hoja.append(dato)

        # Guardar el libro de trabajo
        libro.save("profesores.xlsx")
        ruta_archivo = os.path.abspath("profesores.xlsx")
        os.startfile(ruta_archivo)

    def exportar_a_excel_alumnos():
        libro = openpyxl.Workbook()
        hoja = libro.active

        # Agregar encabezados de columna
        encabezados = ["Nombre", "Apellido", "Usuario", "Contraseña", "Profesor"]
        hoja.append(encabezados)

        # Obtener datos de la tabla de alumnos
        datos = []
        for item in tree_alumnos.get_children():
            valores = tree_alumnos.item(item, 'values')
            datos.append(valores)

        # Agregar los datos a la hoja de Excel
        for dato in datos:
            hoja.append(dato)

        # Guardar el libro de trabajo
        libro.save("alumnos.xlsx")
        ruta_archivo = os.path.abspath("alumnos.xlsx")
        os.startfile(ruta_archivo)

    def exportar_a_excel_calificaciones():
        libro = openpyxl.Workbook()
        hoja = libro.active

        # Agregar encabezados de columna
        encabezados = ["Nombre", "Primer Examen", "Segundo Examen", "Tercer Examen", "Examen Final", "Promedio"]
        hoja.append(encabezados)

        # Obtener datos de la tabla de calificaciones
        datos = []
        for item in tree_calificaciones.get_children():
            valores = tree_calificaciones.item(item, 'values')
            datos.append(valores)

        # Agregar los datos a la hoja de Excel
        for dato in datos:
            hoja.append(dato)

        # Guardar el libro de trabajo
        libro.save("calificaciones.xlsx")
        ruta_archivo = os.path.abspath("calificaciones.xlsx")
        os.startfile(ruta_archivo)

    def eliminar_profesor_seleccionado():
        seleccionado = tree.selection()
        if not seleccionado:
            messagebox.showinfo("Información", "No hay profesor seleccionado.")
            return

        nombre_profesor = tree.item(seleccionado, "values")[0]
        apellido_profesor = tree.item(seleccionado, "values")[1]
        respuesta = messagebox.askyesno("Confirmación", f"¿Desea eliminar a {nombre_profesor} {apellido_profesor}?")

        if respuesta:
            cursor = connection.cursor()
            try:
                # Obtenemos el nombre del profesor desde la fila seleccionada
                nombre_profesor = tree.item(seleccionado, "values")[0]

                # Ejecutamos la consulta para eliminar el profesor
                cursor.execute("DELETE FROM profesores WHERE nombre = ?", (nombre_profesor,))
                connection.commit()

                # Eliminamos la fila seleccionada de la tabla
                tree.delete(seleccionado)

                messagebox.showinfo("Éxito", f"Profesor {nombre_profesor} eliminado correctamente.")
                obtener_cantidad_profesores()
            except Exception as e:
                connection.rollback()
                messagebox.showerror("Error", f"No se pudo eliminar al profesor: {str(e)}")

    def obtener_cantidad_profesores():
        try:
            # Establecer la conexión con la base de datos
            cursor = connection.cursor()
            
            # Consulta SQL para obtener la cantidad de profesores
            query = "SELECT COUNT(*) FROM profesores"
            cursor.execute(query)
            
            # Obtener el resultado y retornar la cantidad de profesores
            cantidad_profesores = cursor.fetchone()[0]
            return cantidad_profesores
            
        except pyodbc.Error as e:
            print("Error en la conexión a la base de datos:", e)
            return 0  # Si hay un error, retornar 0 profesores

    def obtener_cantidad_estudiantes():
        try:
            # Establecer la conexión con la base de datos
            cursor = connection.cursor()
            
            # Consulta SQL para obtener la cantidad de estudiantes
            query = "SELECT COUNT(*) FROM alumnos"
            cursor.execute(query)
            
            # Obtener el resultado y retornar la cantidad de estudiantes
            cantidad_estudiantes = cursor.fetchone()[0]
            return cantidad_estudiantes
            
        except pyodbc.Error as e:
            print("Error en la conexión a la base de datos:", e)
            return 0  # Si hay un error, retornar 0 estudiantes

    def obtener_promedio_calificaciones():
        try:
            # Establecer la conexión con la base de datos
            cursor = connection.cursor()
            
            # Consulta SQL para obtener el promedio de calificaciones
            query = "SELECT AVG(Primer_examen + Segundo_examen + Tercer_examen + Examen_final) / 4 FROM calificaciones"
            cursor.execute(query)
            
            # Obtener el resultado y retornar el promedio de calificaciones
            promedio_calificaciones = cursor.fetchone()[0]
            return promedio_calificaciones
            
        except pyodbc.Error as e:
            print("Error en la conexión a la base de datos:", e)
            return None


    ventana_admin = tk.Toplevel()
    ventana_admin.title('Ventana Administradores')
    ventana_admin.geometry('1250x700')
    ventana_admin.configure(bg="#fff")
    ventana_admin.resizable(False, False)
    utl.centrar_ventana(ventana_admin, 1250, 700)

    # Crear los frames para dividir la ventana
    #frame1 = Frame(ventana_admin, bg='white', highlightthickness=1, highlightbackground="black")
    frame1 = Frame(ventana_admin, bg='white')
    frame2 = Frame(ventana_admin, bg='white')
    frame3 = Frame(ventana_admin, bg='white')

    # Configurar dimensiones y posiciones de los frames usando el método place
    frame1.place(x=10, y=10, width=700, height=380)
    frame2.place(x=720, y=10, width=570, height=380)
    frame3.place(x=10, y=400, width=1230, height=280)

    #cerrar sesion

    def cerrar_sesion():
        respuesta = messagebox.askyesno("Cerrar Sesión", "¿Estás seguro que quieres cerrar la sesión?")
        if respuesta:
            ventana_admin.destroy() 

    img2 = tk.PhotoImage(file='source/img/apagar.png')
    boton_cerrar_sesion = Label(ventana_admin, image=img2, bg='white', cursor='hand2')
    boton_cerrar_sesion.place(x=1200, y=5)
    boton_cerrar_sesion.bind("<Button-1>", lambda e: cerrar_sesion())

    #Hora y fecha

    def update_clock(label):
        locale.setlocale(locale.LC_TIME, "es_ES")
        now = time.strftime("%H:%M:%S")
        label.configure(text=now)
        label.after(1000, lambda: update_clock(label))

    def get_date(label):
        datetime_object = datetime.datetime.now()
        week_day = datetime_object.strftime("%A")

        today = date.today()
        d1 = week_day + ',' + today.strftime("%d/%m/%Y")
        label.configure( text=d1 )

    current_day_label = Label(frame1, text="", font=('Helvetica',20,'bold'), fg='#57a1f8', bg='white')
    current_day_label.pack(padx=10, pady=5)
    current_day_label.place(x=422, y=5)

    current_time_label = Label(frame1, text="", font=('Helvetica',20,'bold'), fg='#57a1f8', bg='white')
    current_time_label.pack(padx=10, pady=5)
    current_time_label.place(x=585, y=40)

    update_clock(current_time_label)
    get_date(current_day_label)

    #circular
    cantidad_profesores = obtener_cantidad_profesores()  # Debes implementar esta función
    cantidad_estudiantes = obtener_cantidad_estudiantes()
    promedio_calificaciones = obtener_promedio_calificaciones()

    def getColorPromedio(p):
        if p > 90:
            return '#00CC27'
        if p > 80:
            return '#ffc107'
        else:
            return '#dc3545'

    canvas1 = Canvas(frame1, bg='white')
    canvas1.place(x=10, y=135, width=200, height=200)
    texto1 = canvas1.create_text(100,90, text=str(cantidad_profesores), font=('Helvetica', 44, 'bold'))
    texto_canvas1 = canvas1.create_text(100, 130, text='Profesores', font=('Helvetica', 14, 'bold'))
    circulo1 = canvas1.create_oval(10, 10, 190, 190, outline='#57a1f8', width=8)

    canvas2 = Canvas(frame1, bg='white')
    canvas2.place(x=255, y=135, width=200, height=200)
    texto2 = canvas2.create_text(100,90, text=str(cantidad_estudiantes), font=('Helvetica', 44, 'bold'))
    texto_canvas2 = canvas2.create_text(100, 130, text='Estudiantes', font=('Helvetica', 14, 'bold'))
    circulo2 = canvas2.create_oval(10, 10, 190, 190, outline='#57a1f8', width=8)

    canvas3 = Canvas(frame1, bg='white')
    canvas3.place(x=500, y=135, width=200, height=200)
    texto3 = canvas3.create_text(100,90, text=str(promedio_calificaciones), font=('Helvetica', 44, 'bold'))
    texto_canvas3 = canvas3.create_text(100, 130, text='Promedio', font=('Helvetica', 14, 'bold'))
    circulo3 = canvas3.create_oval(10, 10, 190, 190, outline=getColorPromedio(promedio_calificaciones), width=8)

    #Image principal
    img = PhotoImage(file='source/img/admin.png')
    Label(frame2, image=img, bg='white').place(x=35,y=15)

    # Crear el notebook y sus pestañas
    notebook = ttk.Notebook(frame3)
    tab1 = Frame(notebook)
    tab2 = Frame(notebook)
    tab3 = Frame(notebook)
    notebook.add(tab1, text='Profesores')
    notebook.add(tab2, text='Estudiantes')
    notebook.add(tab3, text='Calificaciones')
    notebook.place(x=10, y=10)  # Establecer la posición del notebook dentro de frame3

    #Label(tab1, text='Estos son los Profesores en el tab 1', width=164, height=15, bg='white').pack()
    # Tabla para mostrar profesores y cantidad de estudiantes
    tree = ttk.Treeview(tab1, columns=("Nombre", "Apellido", "Usuario", "Contraseña", "Cantidad Estudiantes"), show="headings")
    tree.heading("Nombre", text="Nombre")
    tree.heading("Apellido", text="Apellido")
    tree.heading("Usuario", text="Usuario")
    tree.heading("Contraseña", text="Contraseña")
    tree.heading("Cantidad Estudiantes", text="Cantidad Estudiantes")

    tree.column("Nombre", width=240)
    tree.column("Apellido", width=240)
    tree.column("Usuario", width=240)
    tree.column("Contraseña", width=240)
    tree.column("Cantidad Estudiantes", width=240)
    tree.pack()
    cargar_datos()

    img_excel_tab1 = PhotoImage(file='source/img/sheets.png')
    export_excel_tab1 = Button(tab1, image=img_excel_tab1, bd=0, bg='white', cursor='hand2', command=exportar_a_excel_profesores).place(x=1165,y=185)

    btneliminar = Button(frame1, width=30, pady=6, text='Eliminar Profesor', bg='#dc3545', fg='white', border=0, cursor='hand2', command=lambda:eliminar_profesor_seleccionado())
    btneliminar.place(x=10, y=350)
    btrefres = Button(frame1, width=30, pady=6, text='Refrescar', bg='#ffc107', fg='white', border=0, cursor='hand2', command=lambda:cargar_datos())
    btrefres.place(x=245, y=350)
    btnNuevo = Button(frame1, width=33, pady=6, text='Agregar Nuevo Profesor', bg='#57a1f8', fg='white', border=0, cursor='hand2', command=lambda:main.ventana_register())
    btnNuevo.place(x=480, y=350)

    # Tabla para mostrar alumnos y su profesor
    tree_alumnos = ttk.Treeview(tab2, columns=("Nombre", "Apellido", "Usuario", "Contraseña", "Profesor"), show="headings")
    tree_alumnos.heading("Nombre", text="Nombre")
    tree_alumnos.heading("Apellido", text="Apellido")
    tree_alumnos.heading("Usuario", text="Usuario")
    tree_alumnos.heading("Contraseña", text="Contraseña")
    tree_alumnos.heading("Profesor", text="Profesor")

    tree_alumnos.column("Nombre",width=240)
    tree_alumnos.column("Apellido", width=240)
    tree_alumnos.column("Usuario",width=240)
    tree_alumnos.column("Contraseña",width=240)
    tree_alumnos.column("Profesor", width=240)
    tree_alumnos.pack()
    cargar_datos_alumnos()

    img_excel_tab2 = PhotoImage(file='source/img/sheets.png')
    export_excel_tab2 = Button(tab2, image=img_excel_tab2, bd=0, bg='white', cursor='hand2', command=exportar_a_excel_alumnos)
    export_excel_tab2.place(x=1165,y=185)

    # Tabla para mostrar calificaciones y nombre/apellido de estudiantes
    tree_calificaciones = ttk.Treeview(tab3, columns=("Nombre", "Primer Examen", "Segundo Examen", "Tercer Examen", "Examen Final", "Promedio"), show="headings")
    tree_calificaciones.heading("Nombre", text="Nombre")
    tree_calificaciones.heading("Primer Examen", text="Primer Examen")
    tree_calificaciones.heading("Segundo Examen", text="Segundo Examen")
    tree_calificaciones.heading("Tercer Examen", text="Tercer Examen")
    tree_calificaciones.heading("Examen Final", text="Examen Final")
    tree_calificaciones.heading("Promedio", text="Promedio")
    tree_calificaciones.pack()
    # Asignar el ancho exacto a cada columna (en píxeles)
    tree_calificaciones.column("Nombre", width=200)
    tree_calificaciones.column("Primer Examen", width=200 , anchor="center")
    tree_calificaciones.column("Segundo Examen", width=200 , anchor="center")
    tree_calificaciones.column("Tercer Examen", width=200 , anchor="center")
    tree_calificaciones.column("Examen Final", width=200 , anchor="center")
    tree_calificaciones.column("Promedio", width=200 , anchor="center")
    # Empaquetar la tabla
    cargar_datos_calificaciones()

    img_excel_tab3 = PhotoImage(file='source/img/sheets.png')
    export_excel_tab3 = Button(tab3, image=img_excel_tab2, bd=0, bg='white', cursor='hand2', command=exportar_a_excel_calificaciones)
    export_excel_tab3.place(x=1165,y=185)

    # Configurar pesos para que frame3 ocupe el espacio restante
    ventana_admin.grid_rowconfigure(0, weight=1)
    ventana_admin.grid_rowconfigure(1, weight=1)
    ventana_admin.grid_columnconfigure(0, weight=1)
    ventana_admin.grid_columnconfigure(1, weight=1)

    ventana_admin.mainloop()